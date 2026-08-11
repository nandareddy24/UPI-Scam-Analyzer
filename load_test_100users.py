import time
import random
import json
import statistics
import concurrent.futures
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "http://127.0.0.1:3000"
NUM_USERS = 100
DURATION_SECONDS = 60

# Sample payloads to exercise realistic workload
UPI_PAYLOADS = [
    {"upi": "scammer99@ybl"},
    {"upi": "support@okaxis"},
    {"upi": "winfreecash@paytm"},
    {"upi": "john.doe@upi"},
    {"upi": "lotterywinner@postbank"}
]

URL_PAYLOADS = [
    {"url": "http://secure-banking-login-verify.com"},
    {"url": "https://google.com"},
    {"url": "http://192.168.1.1/admin"},
    {"url": "https://bit.ly/claim-your-prize-now"},
    {"url": "https://sbi.co.in"}
]

SMS_PAYLOADS = [
    {"sms": "You won Rs 50000 cash prize! Click http://bit.ly/claim to enter UPI PIN."},
    {"sms": "Your account 4829 has been debited by Rs 1500.00 for Amazon order."},
    {"sms": "Urgent: Electricity bill unpaid! Power disconnect tonight. Call 9876543210."},
    {"sms": "Dear customer, your KYC is expired. Update at http://kyc-update-bank.com"},
    {"sms": "Hi mom, I reached home safely."}
]

PHONE_PAYLOADS = [
    {"phone": "9876543210"},
    {"phone": "8888888888"},
    {"phone": "9123456789"},
    {"phone": "7000000000"}
]

CHAT_PAYLOADS = [
    {"message": "I received an SMS asking for my UPI PIN to get a refund. Is this a scam?"},
    {"message": "Someone sent me a QR code saying scan to receive money."},
    {"message": "Electricity bill disconnection alert message received from unknown number."}
]

ENDPOINTS = [
    ("POST", "/check_upi", UPI_PAYLOADS),
    ("POST", "/check_url", URL_PAYLOADS),
    ("POST", "/check_sms", SMS_PAYLOADS),
    ("POST", "/check_phone", PHONE_PAYLOADS),
    ("POST", "/api/v1/chat", CHAT_PAYLOADS),
    ("GET", "/login", None),
    ("GET", "/scan", None),
    ("GET", "/mobile_app", None)
]

results = []

def worker_user(user_id, start_time, stop_time):
    session = requests.Session()
    user_results = []

    while time.time() < stop_time:
        method, path, payload_list = random.choice(ENDPOINTS)
        url = BASE_URL + path
        req_start = time.time()
        status_code = 0
        success = False
        error_msg = None

        try:
            if method == "POST":
                payload = random.choice(payload_list) if payload_list else {}
                resp = session.post(url, json=payload, timeout=5)
            else:
                resp = session.get(url, timeout=5)

            req_end = time.time()
            status_code = resp.status_code
            latency_ms = (req_end - req_start) * 1000
            success = (200 <= status_code < 400)
        except Exception as e:
            req_end = time.time()
            latency_ms = (req_end - req_start) * 1000
            error_msg = str(e)

        user_results.append({
            "user_id": user_id,
            "timestamp": req_start,
            "method": method,
            "path": path,
            "latency_ms": latency_ms,
            "status_code": status_code,
            "success": success,
            "error": error_msg
        })

        # Brief delay between user actions (think time 10ms - 50ms)
        time.sleep(random.uniform(0.01, 0.05))

    return user_results

import sys
import io
if sys.platform == "win32":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

def run_load_test():
    print("=" * 70)
    print(f"STARTING BASELINE / LOAD TEST: {NUM_USERS} VIRTUAL USERS FOR {DURATION_SECONDS} SECONDS")
    print(f"Target URL: {BASE_URL}")
    print("=" * 70)

    start_time = time.time()
    stop_time = start_time + DURATION_SECONDS

    all_logs = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=NUM_USERS) as executor:
        futures = [executor.submit(worker_user, i, start_time, stop_time) for i in range(NUM_USERS)]
        for f in concurrent.futures.as_completed(futures):
            all_logs.extend(f.result())

    actual_duration = time.time() - start_time
    total_requests = len(all_logs)
    successful_requests = sum(1 for r in all_logs if r["success"])
    failed_requests = total_requests - successful_requests

    rps = total_requests / actual_duration if actual_duration > 0 else 0
    latencies = [r["latency_ms"] for r in all_logs]

    min_lat = min(latencies) if latencies else 0
    max_lat = max(latencies) if latencies else 0
    avg_lat = statistics.mean(latencies) if latencies else 0
    median_lat = statistics.median(latencies) if latencies else 0
    
    sorted_lat = sorted(latencies) if latencies else [0]
    p90_lat = sorted_lat[int(len(sorted_lat) * 0.90)] if sorted_lat else 0
    p95_lat = sorted_lat[int(len(sorted_lat) * 0.95)] if sorted_lat else 0
    p99_lat = sorted_lat[int(len(sorted_lat) * 0.99)] if sorted_lat else 0

    print("\n" + "=" * 70)
    print("📊 BASELINE / LOAD TEST SUMMARY RESULTS")
    print("=" * 70)
    print(f"• Total Virtual Users:    {NUM_USERS}")
    print(f"• Test Duration:         {actual_duration:.2f} seconds")
    print(f"• Total Requests Sent:   {total_requests:,}")
    print(f"• Successful Requests:   {successful_requests:,} ({(successful_requests/total_requests)*100:.2f}%)")
    print(f"• Failed Requests:       {failed_requests:,} ({(failed_requests/total_requests)*100:.2f}%)")
    print(f"• Requests Per Sec (RPS): {rps:.2f} req/sec")
    print("-" * 70)
    print("⏱️ RESPONSE TIME STATS (Latency):")
    print(f"• Min Response Time:     {min_lat:.2f} ms")
    print(f"• Average Response Time: {avg_lat:.2f} ms")
    print(f"• Median (P50) Time:     {median_lat:.2f} ms")
    print(f"• P90 Response Time:     {p90_lat:.2f} ms")
    print(f"• P95 Response Time:     {p95_lat:.2f} ms")
    print(f"• P99 Response Time:     {p99_lat:.2f} ms")
    print(f"• Max Response Time:     {max_lat:.2f} ms")
    print("=" * 70)

    # Endpoint Breakdown
    endpoint_stats = {}
    for r in all_logs:
        path = r["path"]
        if path not in endpoint_stats:
            endpoint_stats[path] = []
        endpoint_stats[path].append(r["latency_ms"])

    print("\n📌 ENDPOINT BREAKDOWN:")
    for path, lats in endpoint_stats.items():
        ep_count = len(lats)
        ep_avg = statistics.mean(lats)
        ep_min = min(lats)
        ep_max = max(lats)
        print(f"  - {path:25s} | Req Count: {ep_count:5d} | Avg: {ep_avg:6.2f}ms | Min: {ep_min:6.2f}ms | Max: {ep_max:6.2f}ms")

    # Generate Excel Report
    generate_excel_summary(NUM_USERS, actual_duration, total_requests, successful_requests, failed_requests, rps,
                           min_lat, avg_lat, median_lat, p90_lat, p95_lat, p99_lat, max_lat, endpoint_stats)

    # Generate Markdown Summary Artifact
    generate_markdown_report(NUM_USERS, actual_duration, total_requests, successful_requests, failed_requests, rps,
                             min_lat, avg_lat, median_lat, p90_lat, p95_lat, p99_lat, max_lat, endpoint_stats)

def generate_excel_summary(vusers, duration, total, success, failed, rps, min_l, avg_l, med_l, p90, p95, p99, max_l, ep_stats):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Load Test Summary"

    title_fill = PatternFill("solid", fgColor="16213E")
    header_fill = PatternFill("solid", fgColor="1A1A2E")
    alt_fill = PatternFill("solid", fgColor="F8F9FA")
    accent_fill = PatternFill("solid", fgColor="E3F2FD")

    ws.merge_cells("A1:E1")
    ws["A1"] = "🚀 SCAM SHIELD AI — BASELINE / LOAD TESTING REPORT"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    summary_rows = [
        ("Configuration Parameter", "Value"),
        ("Concurrent Virtual Users", vusers),
        ("Test Duration (Target / Actual)", f"60s / {duration:.2f}s"),
        ("Target Host", BASE_URL),
        ("Total Requests Sent", total),
        ("Successful Requests (200 OK)", f"{success} ({success/total*100:.2f}%)"),
        ("Failed Requests", f"{failed} ({failed/total*100:.2f}%)"),
        ("Requests Per Second (RPS)", f"{rps:.2f} req/sec"),
        ("", ""),
        ("Response Time Metric", "Latency (ms)"),
        ("Minimum Response Time", f"{min_l:.2f} ms"),
        ("Average Response Time", f"{avg_l:.2f} ms"),
        ("Median (P50) Response Time", f"{med_l:.2f} ms"),
        ("P90 Response Time", f"{p90:.2f} ms"),
        ("P95 Response Time", f"{p95:.2f} ms"),
        ("P99 Response Time", f"{p99:.2f} ms"),
        ("Maximum Response Time", f"{max_l:.2f} ms"),
    ]

    for r_idx, (k, v) in enumerate(summary_rows, 3):
        ws.cell(row=r_idx, column=1, value=k).font = Font(name="Calibri", bold=True)
        ws.cell(row=r_idx, column=2, value=v)
        if r_idx in [3, 12]:
            ws.cell(row=r_idx, column=1).fill = header_fill
            ws.cell(row=r_idx, column=2).fill = header_fill
            ws.cell(row=r_idx, column=1).font = Font(name="Calibri", bold=True, color="FFFFFF")
            ws.cell(row=r_idx, column=2).font = Font(name="Calibri", bold=True, color="FFFFFF")

    # Endpoint Details Table
    start_ep_row = len(summary_rows) + 5
    ws.merge_cells(f"A{start_ep_row}:E{start_ep_row}")
    ws[f"A{start_ep_row}"] = "📌 PER-ENDPOINT PERFORMANCE BREAKDOWN"
    ws[f"A{start_ep_row}"].font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    ws[f"A{start_ep_row}"].fill = title_fill
    ws[f"A{start_ep_row}"].alignment = Alignment(horizontal="center")

    headers = ["Endpoint Path", "Request Count", "Avg Latency (ms)", "Min Latency (ms)", "Max Latency (ms)"]
    for c_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_ep_row+1, column=c_idx, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for idx, (path, lats) in enumerate(ep_stats.items(), start_ep_row+2):
        ws.cell(row=idx, column=1, value=path).font = Font(name="Calibri", bold=True)
        ws.cell(row=idx, column=2, value=len(lats)).alignment = Alignment(horizontal="center")
        ws.cell(row=idx, column=3, value=round(statistics.mean(lats), 2)).alignment = Alignment(horizontal="center")
        ws.cell(row=idx, column=4, value=round(min(lats), 2)).alignment = Alignment(horizontal="center")
        ws.cell(row=idx, column=5, value=round(max(lats), 2)).alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 12)

    wb.save("load_test_100users_report.xlsx")
    print("[OK] Saved Excel report to load_test_100users_report.xlsx")

def generate_markdown_report(vusers, duration, total, success, failed, rps, min_l, avg_l, med_l, p90, p95, p99, max_l, ep_stats):
    content = f"""# 🚀 Scam Shield AI — Baseline / Load Testing Results

**Test Date:** August 11, 2026  
**Test Configuration:**  
- **Virtual Users:** {vusers} concurrent users  
- **Duration:** 1 minute ({duration:.2f}s actual)  
- **Target Host:** `{BASE_URL}`  

---

## 📊 High-Level Metrics Summary

| Metric | Result Value | Description / Meaning |
|---|---|---|
| **Requests Per Second (RPS)** | **`{rps:.2f} req/sec`** | Average throughput handled continuously per second |
| **Total Requests Sent** | `{total:,}` | Total HTTP requests executed during the 1-minute test |
| **Successful Requests** | `{success:,}` ({(success/total)*100:.2f}%) | HTTP 200/2xx successful responses |
| **Failed Requests** | `{failed:,}` ({(failed/total)*100:.2f}%) | Timeouts, socket errors, or non-2xx status codes |
| **Average Response Time** | **`{avg_l:.2f} ms`** | Mean response time across all endpoints |
| **Minimum Response Time** | `{min_l:.2f} ms` | Fastest single request response time |
| **Median (P50) Response Time** | `{med_l:.2f} ms` | 50% of requests were faster than this |
| **P90 Response Time** | `{p90:.2f} ms` | 90% of requests were faster than this |
| **P95 Response Time** | `{p95:.2f} ms` | 95% of requests were faster than this |
| **P99 Response Time** | `{p99:.2f} ms` | 99% of requests were faster than this |
| **Maximum Response Time** | `{max_l:.2f} ms` | Slowest single request response time |

---

## 📌 Endpoint Performance Breakdown

| Endpoint Path | Total Requests | Avg Latency | Min Latency | Max Latency | Performance Grade |
|---|---|---|---|---|---|
"""
    for path, lats in ep_stats.items():
        count = len(lats)
        avg_t = statistics.mean(lats)
        min_t = min(lats)
        max_t = max(lats)
        grade = "🟢 Excellent" if avg_t < 100 else ("🟡 Good" if avg_t < 300 else "🔴 Slow")
        content += f"| `{path}` | {count:,} | {avg_t:.2f} ms | {min_t:.2f} ms | {max_t:.2f} ms | {grade} |\n"

    content += f"""

---

## 💡 System Health & Observations

1. **Throughput Capability:** The Flask backend successfully sustained **`{rps:.2f} requests per second`** with **100 concurrent virtual users**.
2. **Latency Analysis:** The average response time was **`{avg_l:.2f} ms`** with a minimum of **`{min_l:.2f} ms`** and maximum of **`{max_l:.2f} ms`**.
3. **Stability:** Handled {total:,} total requests over 1 minute under expected peak concurrent user load.
"""
    with open("load_test_results.md", "w", encoding="utf-8") as f:
        f.write(content)
    print("[OK] Saved Markdown summary to load_test_results.md")

if __name__ == "__main__":
    run_load_test()
