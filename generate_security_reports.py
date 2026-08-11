"""
Security Assessment Excel Report Generator
Scam Shield AI -- UPI-Scam-Analyzer
Generates: findings.xlsx, endpoint-inventory.xlsx, dependency-report.xlsx
Run from project root: python generate_security_reports.py
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import datetime, os

COLORS = {
    "critical": "C0392B", "high": "E67E22", "medium": "F1C40F",
    "low": "3498DB", "info": "95A5A6", "safe": "27AE60",
    "header_bg": "1A1A2E", "header_fg": "FFFFFF",
    "row_alt": "F8F9FA", "row_main": "FFFFFF",
    "border": "DEE2E6", "title_bg": "16213E",
}

def make_border():
    thin = Side(style='thin', color=COLORS["border"])
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_font(size=11):
    return Font(name='Calibri', bold=True, color="FFFFFF", size=size)

def cell_font(size=10, bold=False, color="000000"):
    return Font(name='Calibri', bold=bold, size=size, color=color)

def severity_fill(sev):
    mapping = {"Critical": COLORS["critical"], "High": COLORS["high"],
               "Medium": COLORS["medium"], "Low": COLORS["low"], "Info": COLORS["info"]}
    return make_fill(mapping.get(sev, "CCCCCC"))

def severity_fg(sev):
    return "FFFFFF" if sev in ("Critical", "High", "Info") else "000000"

FINDINGS = [
    (1, "Critical", "A02:2021", "CWE-798", "Hardcoded Secret Key Fallback", "app.py", "29", "ALL routes",
     "Flask uses 'secret123' fallback. Attacker can forge session cookies, impersonate any user including admin.",
     "Require SECRET_KEY via os.getenv(); raise RuntimeError if not set."),
    (2, "Critical", "A01:2021", "CWE-798", "Hardcoded Admin Email in Source", "app.py", "31", "ALL admin routes",
     "ADMIN_EMAIL='nandakumarreddy63@gmail.com' hardcoded. Visible to repo access; admin identity leaked in API responses.",
     "Move ADMIN_EMAIL to environment variable. Raise RuntimeError if missing."),
    (3, "Critical", "A01:2021", "CWE-306", "Missing Auth on Core Scan APIs", "app.py", "1203-1680",
     "/check_upi, /check_url, /check_sms, /check_phone, /api/v1/analyze, /api/v1/chat",
     "All scan APIs accessible without login. Enables mass abuse, API quota drain, unauthorized DB writes, DoS.",
     "Create @login_required decorator and apply to all scan endpoints."),
    (4, "Critical", "A01:2021", "CWE-434", "Unauthenticated File Upload", "app.py", "1002-1090",
     "POST /scan_qr, POST /scan_ocr",
     "No auth, no size limit, no MIME validation on file upload endpoints. DoS via large files; arbitrary image processing.",
     "Add @login_required, MAX_CONTENT_LENGTH=5MB, MIME type validation."),
    (5, "Critical", "A01:2021", "CWE-639", "IDOR on PDF Report Download", "app.py", "1093-1166",
     "GET /download_pdf_report/<scan_id>",
     "Any authenticated user can download other users' PDFs by incrementing scan_id. Full scan history exfiltration.",
     "Add WHERE user_id=%s filter to the scan query in download_pdf_report."),
    (6, "Critical", "A01:2021", "CWE-285", "Zero Auth on Blacklist Delete", "app.py", "963-973",
     "POST /delete_blacklist/<id>",
     "No auth check at all. Any internet user can delete blacklist entries, unblocking scam UPIs/URLs.",
     "Add session check and ADMIN_EMAIL comparison before executing DELETE."),
    (7, "Critical", "A04:2021", "CWE-362", "In-Memory OTP Race Condition", "app.py", "199-200",
     "POST /verify_otp, POST /verify_reset_otp",
     "OTPs in global Python dict -- not thread-safe, not persistent. Multi-worker Gunicorn loses OTPs across workers.",
     "Use Redis or database table for OTP storage with atomic operations."),
    (8, "High", "A07:2021", "CWE-307", "No Rate Limiting on Auth Endpoints", "app.py", "728-774",
     "POST /login, POST /verify_otp, POST /forgot",
     "No brute force protection. 6-digit OTPs (1M combos) brute-forceable in 10min window. No account lockout.",
     "Install flask-limiter. Apply @limiter.limit('5/minute') to auth endpoints."),
    (9, "High", "A05:2021", "CWE-942", "CORS Wildcard on Extension API", "app.py", "1312-1345",
     "POST /api/v1/extension/check_url",
     "Access-Control-Allow-Origin: * allows any website to call this API, enabling CSRF-like abuse and data exfiltration.",
     "Restrict CORS to specific chrome-extension:// ID via environment variable."),
    (10, "High", "A07:2021", "CWE-640", "Password Reset Auth Bypass", "app.py", "826-854",
     "POST /update_password",
     "update_password accepts email+new password without checking OTP was verified. Any email known = full account takeover.",
     "Track OTP verification state; check verified flag before allowing password update."),
    (11, "High", "A01:2021", "CWE-200", "History Exposes All Users Data", "app.py", "983-988",
     "GET /history",
     "No WHERE user_id filter. Any logged-in user sees all users' scan histories -- UPI handles, phones, URLs, SMS.",
     "Add WHERE user_id=%s using get_current_user_id() in the history query."),
    (12, "High", "A01:2021", "CWE-200", "Export CSV Dumps All Users Data", "app.py", "1639",
     "GET /export_data",
     "Any authenticated user downloads full CSV of all users' scan records -- complete database dump.",
     "Add WHERE user_id=%s filter to export_data query."),
    (13, "High", "A01:2021", "CWE-200", "Dashboard Leaks Global Statistics", "app.py", "876-889",
     "GET /dashboard",
     "Global scan counts and last 5 scans from ALL users shown to any authenticated user. Privacy violation.",
     "Filter all dashboard queries by current user_id."),
    (14, "High", "A05:2021", "CWE-94", "Flask Debug Mode Enabled", "app.py", "1767",
     "ALL routes (Werkzeug debugger)",
     "debug=True exposes interactive Python REPL via Werkzeug debugger on exceptions. Remote code execution risk.",
     "Set debug=os.getenv('FLASK_DEBUG','false').lower()=='true'. Never True in prod."),
    (15, "High", "A07:2021", "CWE-204", "User Enumeration via Error Messages", "app.py", "746-747",
     "POST /login, POST /forgot",
     "Different messages 'User not found' vs 'Invalid Password' allow email enumeration.",
     "Return identical generic error 'Invalid email or password' for all auth failures."),
    (16, "High", "A09:2021", "CWE-532", "Sensitive Data in Console Logs", "app.py", "653,817,770",
     "POST /register, POST /forgot, POST /login",
     "OTP values and password hashes printed to console logs. Cloud log aggregators expose these.",
     "Remove all print(otp) and print(stored_password) statements."),
    (17, "Medium", "A01:2021", "CWE-352", "No CSRF Protection", "app.py, templates/", "All POST forms",
     "POST /update_password, /admin, /delete_blacklist, /verify_otp",
     "State-changing forms lack CSRF tokens. Attackers forge requests from malicious websites.",
     "Install flask-wtf. Use CSRFProtect(app) and include csrf_token in all forms."),
    (18, "Medium", "A03:2021", "CWE-79", "No Input Validation on Report Fields", "app.py", "1607-1628",
     "POST /report_scam",
     "reason and proof fields stored without length limits or HTML sanitization. Stored XSS risk in admin panel.",
     "Add max length 500 chars. Use html.escape() or ensure Jinja2 auto-escaping in templates."),
    (19, "Medium", "A05:2021", "CWE-209", "Raw Exception in API Responses", "app.py", "557-558",
     "POST /api/v1/auth/register and others",
     "str(e) returned in JSON responses exposes DB schema, file paths, SQL error details.",
     "Return generic error messages. Log full exception server-side only."),
    (20, "Medium", "A05:2021", "CWE-693", "Missing HTTP Security Headers", "app.py", "N/A",
     "ALL routes",
     "No Content-Security-Policy, X-Frame-Options, X-Content-Type-Options, HSTS, Referrer-Policy.",
     "Add @app.after_request hook setting all security headers."),
    (21, "Medium", "A05:2021", "CWE-770", "No File Upload Size Limit", "app.py", "1011,1041",
     "POST /scan_qr, POST /scan_ocr",
     "Files read entirely into memory. Large uploads (100MB+) exhaust RAM and crash the application.",
     "Set app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024 (5MB)."),
    (22, "Medium", "A03:2021", "CWE-89", "SQL Wildcard Injection in Blacklist", "app.py", "233-260",
     "Internally called by all scan endpoints",
     "User input with % or _ passed unescaped into LIKE queries. Can cause all blacklist entries to match or none.",
     "Escape SQL wildcards: clean_data.replace('%', r'\\%').replace('_', r'\\_')"),
    (23, "Medium", "A07:2021", "CWE-613", "No Session Timeout", "app.py", "N/A",
     "ALL authenticated routes",
     "Sessions never expire server-side. Stolen session cookie grants permanent access until explicit logout.",
     "Set PERMANENT_SESSION_LIFETIME = timedelta(hours=8) and session.permanent = True."),
    (24, "Medium", "A07:2021", "CWE-614", "Insecure Session Cookie Attributes", "app.py", "N/A",
     "ALL authenticated routes",
     "Session cookies missing Secure, HttpOnly, SameSite attributes. XSS cookie theft and CSRF easier.",
     "Set SESSION_COOKIE_SECURE=True, HTTPONLY=True, SAMESITE='Lax'."),
    (25, "Medium", "A07:2021", "CWE-521", "No Password Complexity Policy", "app.py", "622-625",
     "POST /register, POST /update_password",
     "Single-character passwords accepted. No minimum length or complexity requirements.",
     "Enforce min 8 chars, at least 1 number, 1 uppercase."),
    (26, "Medium", "A03:2021", "CWE-20", "No Email Format Validation", "app.py", "622-623",
     "POST /register, POST /api/v1/auth/register",
     "Any string accepted as email. Malformed emails cause OTP failures and potential XSS in admin views.",
     "Validate with regex: r'^[\\w.+-]+@[\\w-]+\\.[\\w.-]+$'"),
    (27, "Medium", "A08:2021", "CWE-502", "Joblib Pickle Deserialization Risk", "app.py", "179-196",
     "Application startup -- all routes affected",
     "joblib.load() uses Python pickle. Malicious .pkl executes arbitrary code at startup.",
     "Verify .pkl file checksums (SHA-256) before loading. Store models outside repository."),
    (28, "Medium", "A01:2021", "CWE-306", "No Auth on Developer Analyze API", "app.py", "1658-1680",
     "POST /api/v1/analyze",
     "Developer API has no authentication, no API key, no rate limiting. Unlimited free fraud detection abuse.",
     "Require API key header or session auth. Apply rate limiting."),
    (29, "Low", "A09:2021", "CWE-532", "Login Error Stack Trace in Logs", "app.py", "770-772",
     "POST /login",
     "Full exception and stored password hash printed to logs on bcrypt errors.",
     "Remove print(stored_password) and print(e) from login error handler."),
    (30, "Low", "N/A", "CWE-312", ".env File in Repository Root", ".env", "N/A", "N/A",
     ".env file present. If committed to public Git, all secrets (API keys, credentials) exposed.",
     "Verify .env is in .gitignore. Never commit .env to version control."),
    (31, "Low", "N/A", "CWE-312", "Browser Extension PEM Key in Repo", "browser_extension.pem", "N/A", "N/A",
     "1704-byte private key for Chrome extension signing present. Can sign malicious extension versions.",
     "Add browser_extension.pem to .gitignore. Store private keys securely outside repository."),
    (32, "Low", "N/A", "CWE-312", "SQLite Database File in Repository", "database.db", "N/A", "N/A",
     "36KB database file. If committed, exposes all user records, password hashes, blacklist entries.",
     "Add database.db to .gitignore. Never commit database files."),
    (33, "Low", "A08:2021", "CWE-502", "ML Model PKL Files in Repository", "*.pkl files", "N/A", "All startup",
     "url_model.pkl (8.7MB) in repository. Tampered models silently alter fraud detection decisions.",
     "Store models outside repository. Verify SHA-256 checksums before loading."),
    (34, "Low", "A09:2021", "CWE-778", "No Security Audit Logging", "app.py", "N/A", "ALL routes",
     "Failed logins, unauthorized access, OTP failures not logged. Incident response impossible.",
     "Add structured logging using Python logging module for all security-relevant events."),
    (35, "Low", "N/A", "CWE-770", "No Input Length Limits in APIs", "app.py", "1216,1276,1436,1532",
     "POST /check_upi, /check_url, /check_sms, /check_phone",
     "Unlimited-length strings accepted. 1MB SMS causes ML vectorizer CPU exhaustion.",
     "Add: if len(input_data) > 5000: return error 400."),
    (36, "Info", "N/A", "N/A", "reportlab Lazy Import", "app.py", "1106",
     "GET /download_pdf_report/<id>",
     "reportlab imported inside function. Import errors surface only at runtime, not at startup.",
     "Move import to module level. Add reportlab to requirements.txt."),
    (37, "Info", "N/A", "N/A", "reportlab Missing from requirements.txt", "requirements.txt", "N/A",
     "GET /download_pdf_report/<id>",
     "reportlab used in application but absent from requirements.txt. Causes production deployment failure.",
     "Add 'reportlab>=4.0.0' to requirements.txt."),
    (38, "Info", "N/A", "N/A", "Shared DB Cursor Not Thread-Safe", "app.py", "170-177",
     "ALL database queries",
     "Global cursor singleton shared across all requests. Concurrent requests can mix query results.",
     "Use per-request connections or Flask-SQLAlchemy with connection pooling."),
    (39, "Info", "N/A", "CWE-338", "Weak OTP PRNG (random.randint)", "app.py", "640,803",
     "POST /register, POST /forgot",
     "Python random module uses Mersenne Twister -- not cryptographically secure.",
     "Use: import secrets; otp = str(secrets.randbelow(900000) + 100000)"),
    (40, "Info", "N/A", "N/A", "Duplicate Route Versioning", "app.py", "1203-1204",
     "ALL scan endpoints",
     "Routes /check_upi and /api/v1/scan/upi both map to same handler. No versioning strategy.",
     "Deprecate unversioned routes. Direct all traffic to /api/v1/* routes only."),
    (41, "Info", "N/A", "N/A", "No Content-Type Validation on JSON", "app.py", "Multiple",
     "ALL JSON API endpoints",
     "request.get_json(silent=True) silently returns None on malformed Content-Type.",
     "Validate request.content_type == 'application/json' before processing."),
]

ENDPOINTS = [
    (1, "GET", "/", "No", "None", "HTML Redirect", "Redirect to /login", "Low"),
    (2, "GET/POST", "/register", "No", "name, email, password, confirm", "HTML", "OTP page", "Medium"),
    (3, "POST", "/verify_otp", "No", "email, otp (form)", "HTML Redirect", "OTP verification", "High"),
    (4, "POST", "/verify_reset_otp", "No", "email, otp (form)", "HTML", "Reset OTP check", "High"),
    (5, "GET/POST", "/login", "No", "email, password", "HTML Redirect", "Session login", "High"),
    (6, "GET", "/logout", "No", "None", "HTML Redirect", "Session clear", "Low"),
    (7, "GET/POST", "/forgot", "No", "email (form)", "HTML", "Password reset flow", "Medium"),
    (8, "POST", "/update_password", "NONE -- BUG", "email, password, confirm", "HTML Redirect", "CRITICAL: No OTP check before reset", "Critical"),
    (9, "GET", "/dashboard", "Session", "None", "HTML", "HIGH: Shows ALL users' stats", "High"),
    (10, "GET/POST", "/admin", "Session + Admin", "data, type, reason", "HTML", "Admin blacklist management", "High"),
    (11, "POST", "/admin/approve_report/<id>", "Session + Admin", "Path ID", "HTML Redirect", "Approve community report", "Medium"),
    (12, "POST", "/admin/reject_report/<id>", "Session + Admin", "Path ID", "HTML Redirect", "Reject community report", "Medium"),
    (13, "POST", "/delete_blacklist/<id>", "NONE -- BUG", "Path ID", "HTML Redirect", "CRITICAL: No auth -- destroys blacklist", "Critical"),
    (14, "GET", "/scan", "Session", "None", "HTML", "Scan hub page", "Low"),
    (15, "GET", "/history", "Session", "None", "HTML", "HIGH: Shows ALL users' scan history", "High"),
    (16, "GET", "/ocr_qr", "Session", "None", "HTML", "OCR/QR scan page", "Low"),
    (17, "POST", "/scan_qr", "NONE -- BUG", "payload/qr_image (file)", "JSON", "CRITICAL: Unauth file upload", "Critical"),
    (18, "POST", "/scan_ocr", "NONE -- BUG", "screenshot (file)", "JSON", "CRITICAL: Unauth file upload", "Critical"),
    (19, "GET", "/download_pdf_report/<id>", "Session", "Path ID", "PDF Binary", "CRITICAL: IDOR -- no user_id check", "Critical"),
    (20, "GET", "/upi", "Session", "None", "HTML", "UPI scan page", "Low"),
    (21, "GET", "/url", "Session", "None", "HTML", "URL scan page", "Low"),
    (22, "GET", "/sms", "Session", "None", "HTML", "SMS scan page", "Low"),
    (23, "POST", "/check_upi", "NONE -- BUG", "JSON: {upi}", "JSON", "CRITICAL: Unauth UPI scan", "Critical"),
    (24, "POST", "/api/v1/scan/upi", "NONE -- BUG", "JSON: {upi}", "JSON", "CRITICAL: Unauth UPI scan", "Critical"),
    (25, "POST", "/check_phone", "NONE -- BUG", "JSON: {phone}", "JSON", "CRITICAL: Unauth phone scan", "Critical"),
    (26, "POST", "/api/v1/scan/phone", "NONE -- BUG", "JSON: {phone}", "JSON", "CRITICAL: Unauth phone scan", "Critical"),
    (27, "POST", "/check_url", "NONE -- BUG", "JSON: {url}", "JSON", "CRITICAL: Unauth URL scan + VT/SB abuse", "Critical"),
    (28, "POST", "/api/v1/scan/url", "NONE -- BUG", "JSON: {url}", "JSON", "CRITICAL: Unauth URL scan + VT/SB abuse", "Critical"),
    (29, "POST", "/check_sms", "NONE -- BUG", "JSON: {sms}", "JSON", "CRITICAL: Unauth SMS scan", "Critical"),
    (30, "POST", "/api/v1/scan/sms", "NONE -- BUG", "JSON: {sms}", "JSON", "CRITICAL: Unauth SMS scan", "Critical"),
    (31, "GET", "/report", "Session", "None", "HTML", "Fraud report page", "Low"),
    (32, "POST", "/report_scam", "Session", "JSON: {data, type, reason, proof}", "JSON", "Fraud report submission", "Medium"),
    (33, "GET", "/export_data", "Session", "None", "CSV Download", "HIGH: Exports ALL users' data", "High"),
    (34, "POST", "/api/v1/analyze", "NONE -- BUG", "JSON: {type, data}", "JSON", "CRITICAL: Unauth developer API", "Critical"),
    (35, "GET", "/chatbot", "Session", "None", "HTML", "Chatbot page", "Low"),
    (36, "POST", "/api/v1/chat", "NONE -- BUG", "JSON: {message}", "JSON", "HIGH: Unauth chatbot query", "High"),
    (37, "POST/OPTIONS", "/api/v1/extension/check_url", "NONE -- BUG", "JSON: {url}", "JSON + CORS *", "CRITICAL: Unauth + CORS wildcard", "Critical"),
    (38, "GET", "/extension", "Session", "None", "HTML", "Extension page", "Low"),
    (39, "GET", "/download_extension_zip", "Session", "None", "ZIP Download", "Extension download", "Low"),
    (40, "GET", "/mobile", "NONE -- BUG", "None", "HTML", "Mobile app page -- no auth", "Medium"),
    (41, "GET", "/mobile_app", "NONE -- BUG", "None", "HTML", "Mobile app page -- no auth", "Medium"),
    (42, "GET", "/android", "Session", "None", "HTML", "Android page", "Low"),
    (43, "GET", "/download_android_zip", "Session", "None", "ZIP Download", "Android app download", "Low"),
    (44, "POST", "/api/v1/auth/register", "No", "JSON: {name, email, password}", "JSON", "Mobile registration", "Medium"),
    (45, "POST", "/api/v1/auth/login", "No", "JSON: {email, password}", "JSON", "Mobile login", "Medium"),
    (46, "POST", "/api/v1/auth/logout", "No", "None", "JSON", "Mobile logout", "Low"),
    (47, "GET", "/api/v1/auth/me", "Session", "None", "JSON", "Current user info", "Low"),
    (48, "GET", "/api/v1/admin/blacklist", "Session + Admin", "None", "JSON", "Admin blacklist API", "Medium"),
    (49, "GET", "/api/v1/admin/reports", "Session + Admin", "None", "JSON", "Admin reports API", "Medium"),
]

DEPS = [
    ("Flask", ">=3.0.0", "3.0.3", "OK", "Low", "No critical CVEs in 3.x", "Keep updated"),
    ("Flask-Bcrypt", "Latest", "1.0.1", "OK", "Low", "No known CVEs", "Keep updated"),
    ("Flask-Cors", "Latest", "4.0.1", "OK", "Low", "No known CVEs", "Keep updated"),
    ("Flask-Login", "Latest", "0.6.3", "OK", "Low", "No known CVEs", "Keep updated"),
    ("bcrypt", "Latest", "4.1.3", "OK", "Low", "Secure bcrypt implementation", "Keep updated"),
    ("gunicorn", "Latest", "22.0.0", "OK", "Low", "No critical CVEs", "Keep updated"),
    ("psycopg2-binary", "Latest", "2.9.9", "OK", "Low", "No critical CVEs", "Keep updated"),
    ("numpy", "Latest", "1.26.4", "OK", "Low", "No critical CVEs", "Keep updated"),
    ("pandas", "Latest", "2.2.2", "OK", "Low", "No critical CVEs", "Keep updated"),
    ("scikit-learn", "Latest", "1.4.2", "WARN", "Medium", "Joblib pickle deserialization -- model files can contain malicious code", "Verify model checksums"),
    ("scipy", "Latest", "1.13.0", "OK", "Low", "No critical CVEs", "Keep updated"),
    ("joblib", "Latest", "1.4.2", "WARN", "Medium", "Uses pickle internally. Malicious .pkl executes arbitrary code", "Verify model file integrity"),
    ("requests", "Latest", "2.32.3", "OK", "Low", "No critical CVEs in recent versions", "Keep updated"),
    ("python-whois", "Latest", "0.9.4", "OK", "Low", "No known CVEs", "Keep updated"),
    ("python-dotenv", "Latest", "1.0.1", "OK", "Low", "No known CVEs", "Keep updated"),
    ("resend", "Latest", "2.0.0", "OK", "Low", "No known CVEs", "Keep updated"),
    ("opencv-python-headless", "Latest", "4.9.0", "WARN", "Medium", "CVE-2023-4863 affects libwebp in some builds. Verify exact version", "Update to 4.10.0+"),
    ("Pillow", "Latest", "10.3.0", "WARN", "Medium", "CVE-2024-28219 affects <10.3.0. CVE-2023-44271 affects <10.0.1", "Ensure >=10.3.0"),
    ("Jinja2", "Latest", "3.1.4", "OK", "Low", "No critical CVEs in 3.1.x", "Keep updated"),
    ("Werkzeug", "Transitive", "3.0.3", "WARN", "High", "CVE-2023-46136 (DoS) affects <3.0.1. Ensure >=3.0.1", "Verify >=3.0.1"),
    ("reportlab", "NOT LISTED", "Missing", "MISSING", "High", "Used in app.py but absent from requirements.txt. Causes production deployment failure", "Add reportlab>=4.0.0"),
    ("openpyxl", "NOT LISTED", "Installed", "MISSING", "Medium", "Used in Excel report generation but not in requirements.txt", "Add openpyxl>=3.1.0"),
]


def generate_findings_xlsx(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Findings"

    # Title row
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = "SCAM SHIELD AI -- SECURITY ASSESSMENT FINDINGS (41 Vulnerabilities)"
    c.font = Font(name='Calibri', bold=True, size=14, color="FFFFFF")
    c.fill = make_fill(COLORS["title_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.datetime.now().strftime('%B %d, %Y %H:%M')}  |  Overall Risk: CRITICAL  |  Critical: 7  High: 9  Medium: 11  Low: 8  Info: 6"
    c.font = Font(name='Calibri', size=10, color="FFFFFF")
    c.fill = make_fill("0F3460")
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    headers = ["ID", "Severity", "OWASP", "CWE", "Vulnerability Type", "File", "Line", "Endpoint", "Impact", "Fix / Recommendation", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font()
        cell.fill = make_fill(COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = make_border()
    ws.row_dimensions[3].height = 20

    for row_idx, finding in enumerate(FINDINGS, 4):
        id_, sev, owasp, cwe, vtype, file_, line, endpoint, impact, fix = finding
        row_data = [id_, sev, owasp, cwe, vtype, file_, line, endpoint, impact, fix, "Open"]
        bg = COLORS["row_alt"] if row_idx % 2 == 0 else COLORS["row_main"]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = make_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if col_idx == 2:  # Severity badge
                cell.fill = severity_fill(sev)
                cell.font = Font(name='Calibri', bold=True, size=10, color=severity_fg(sev))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 1:  # ID
                cell.font = Font(name='Calibri', bold=True, size=10, color=COLORS["title_bg"])
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = make_fill(bg)
            elif col_idx == 11:  # Status
                cell.fill = make_fill("FFE5E5")
                cell.font = Font(name='Calibri', bold=True, size=10, color=COLORS["critical"])
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.fill = make_fill(bg)
                cell.font = cell_font()

        ws.row_dimensions[row_idx].height = 55

    col_widths = [5, 10, 12, 12, 30, 22, 10, 35, 50, 50, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:K{len(FINDINGS)+3}"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Scam Shield AI -- Security Findings Summary"
    ws2["A1"].font = Font(name='Calibri', bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = make_fill(COLORS["title_bg"])
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 28

    sev_data = [("Critical", 7), ("High", 9), ("Medium", 11), ("Low", 8), ("Informational", 6)]
    for col, h in enumerate(["Severity", "Count", "% of Total", "Status"], 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = header_font()
        cell.fill = make_fill(COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = make_border()

    for row, (sev, count) in enumerate(sev_data, 3):
        sev_key = sev.lower().split()[0]
        color_map = {"critical": COLORS["critical"], "high": COLORS["high"],
                     "medium": COLORS["medium"], "low": COLORS["low"], "informational": COLORS["info"]}
        ws2.cell(row=row, column=1, value=sev).font = Font(name='Calibri', bold=True, size=11, color=color_map.get(sev_key, "000000"))
        ws2.cell(row=row, column=2, value=count).alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=3, value=f"{count/41*100:.1f}%").alignment = Alignment(horizontal="center")
        ws2.cell(row=row, column=4, value="Open").font = Font(name='Calibri', bold=True, color=COLORS["critical"])
        ws2.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        for col in range(1, 5):
            ws2.cell(row=row, column=col).border = make_border()

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 12

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Findings by Severity"
    chart.y_axis.title = "Count"
    data = Reference(ws2, min_col=2, min_row=2, max_row=7)
    cats = Reference(ws2, min_col=1, min_row=3, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 16
    chart.height = 10
    ws2.add_chart(chart, "F2")

    wb.save(filename)
    print(f"[OK] Generated: {filename}")


def generate_endpoint_xlsx(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "API Endpoints"

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "SCAM SHIELD AI -- API ENDPOINT INVENTORY (49 Endpoints)"
    c.font = Font(name='Calibri', bold=True, size=14, color="FFFFFF")
    c.fill = make_fill(COLORS["title_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.datetime.now().strftime('%B %d, %Y')}  |  Total: 49  |  Auth Issues: 18  |  Critical Risk: 12"
    c.font = Font(name='Calibri', size=10, color="FFFFFF")
    c.fill = make_fill("0F3460")
    c.alignment = Alignment(horizontal="center")

    headers = ["#", "Method", "Endpoint", "Auth Required", "Input Parameters", "Response Type", "Security Notes", "Risk Level"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font()
        cell.fill = make_fill(COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = make_border()
    ws.row_dimensions[3].height = 20

    for row_idx, ep in enumerate(ENDPOINTS, 4):
        num, method, path, auth, params, resp_type, notes, risk = ep
        row_data = [num, method, path, auth, params, resp_type, notes, risk]
        bg = COLORS["row_main"] if row_idx % 2 == 0 else COLORS["row_alt"]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = make_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = cell_font()

            if col_idx == 4:  # Auth
                if "BUG" in str(val) or "NONE" in str(val):
                    cell.fill = make_fill("FFE0E0")
                    cell.font = Font(name='Calibri', bold=True, size=10, color=COLORS["critical"])
                else:
                    cell.fill = make_fill("E8F5E9")
                    cell.font = Font(name='Calibri', size=10, color="1B5E20")
            elif col_idx == 8:  # Risk
                risk_map = {
                    "Critical": (COLORS["critical"], "FFFFFF"), "High": (COLORS["high"], "FFFFFF"),
                    "Medium": (COLORS["medium"], "000000"), "Low": (COLORS["low"], "FFFFFF"),
                }
                bg_c, fg_c = risk_map.get(val, ("CCCCCC", "000000"))
                cell.fill = make_fill(bg_c)
                cell.font = Font(name='Calibri', bold=True, size=10, color=fg_c)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 2:  # Method
                method_colors = {"GET": "E8F5E9", "POST": "E3F2FD", "GET/POST": "FFF8E1", "POST/OPTIONS": "F3E5F5"}
                cell.fill = make_fill(method_colors.get(str(val), "F5F5F5"))
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(name='Calibri', bold=True, size=10)
            else:
                cell.fill = make_fill(bg)

        ws.row_dimensions[row_idx].height = 40

    col_widths = [5, 14, 42, 18, 35, 16, 48, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{len(ENDPOINTS)+3}"

    wb.save(filename)
    print(f"[OK] Generated: {filename}")


def generate_dependency_xlsx(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dependencies"

    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "SCAM SHIELD AI -- DEPENDENCY VULNERABILITY REPORT"
    c.font = Font(name='Calibri', bold=True, size=14, color="FFFFFF")
    c.fill = make_fill(COLORS["title_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = f"Generated: {datetime.datetime.now().strftime('%B %d, %Y')}  |  Total: {len(DEPS)} packages  |  Issues: 6  |  Missing from req.txt: 2"
    c.font = Font(name='Calibri', size=10, color="FFFFFF")
    c.fill = make_fill("0F3460")
    c.alignment = Alignment(horizontal="center")

    headers = ["Package", "Version (req.txt)", "Installed Version", "Status", "Risk", "CVE / Issue Details", "Recommendation"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font()
        cell.fill = make_fill(COLORS["header_bg"])
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = make_border()
    ws.row_dimensions[3].height = 20

    for row_idx, dep in enumerate(DEPS, 4):
        pkg, req_ver, inst_ver, status, risk, notes, rec = dep
        row_data = [pkg, req_ver, inst_ver, status, risk, notes, rec]
        bg = COLORS["row_main"] if row_idx % 2 == 0 else COLORS["row_alt"]

        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = make_border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            if col_idx == 4:  # Status
                status_map = {
                    "OK": ("E8F5E9", "1B5E20"),
                    "WARN": ("FFF8E1", "856404"),
                    "MISSING": ("FFE0E0", COLORS["critical"]),
                }
                bg_c, fg_c = status_map.get(str(val), ("FFFFFF", "000000"))
                cell.fill = make_fill(bg_c)
                cell.font = Font(name='Calibri', bold=True, size=10, color=fg_c)
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 5:  # Risk
                risk_map = {
                    "Critical": (COLORS["critical"], "FFFFFF"), "High": (COLORS["high"], "FFFFFF"),
                    "Medium": (COLORS["medium"], "000000"), "Low": (COLORS["low"], "FFFFFF"),
                }
                bg_c, fg_c = risk_map.get(str(val), ("CCCCCC", "000000"))
                cell.fill = make_fill(bg_c)
                cell.font = Font(name='Calibri', bold=True, size=10, color=fg_c)
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 1:
                cell.font = Font(name='Calibri', bold=True, size=10, color=COLORS["title_bg"])
                cell.fill = make_fill("EEF2FF")
            else:
                cell.fill = make_fill(bg)
                cell.font = cell_font()

        ws.row_dimensions[row_idx].height = 45

    col_widths = [25, 18, 18, 12, 10, 58, 35]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:G{len(DEPS)+3}"

    # Scan commands sheet
    ws2 = wb.create_sheet("Scan Commands")
    ws2.merge_cells("A1:B1")
    ws2["A1"].value = "Security Scanning Commands"
    ws2["A1"].font = Font(name='Calibri', bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill = make_fill(COLORS["title_bg"])
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 25

    for col, h in enumerate(["Command Purpose", "Command to Run"], 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = header_font()
        cell.fill = make_fill(COLORS["header_bg"])
        cell.border = make_border()

    commands = [
        ("Dependency Vulnerability Scan", "pip install safety && safety check -r requirements.txt"),
        ("Upgrade All Dependencies", "pip install --upgrade -r requirements.txt"),
        ("Check Outdated Packages", "pip list --outdated"),
        ("Bandit SAST Scan (Python)", "pip install bandit && bandit -r app.py -ll"),
        ("Semgrep Security Scan", "pip install semgrep && semgrep --config=auto app.py"),
        ("Find All Pickle Loads", "grep -rn \"joblib.load\\|pickle.load\" . --include=\"*.py\""),
        ("Check Flask Version", "pip show flask werkzeug"),
        ("Verify Pillow Version", "pip show Pillow"),
    ]
    for row, (name, cmd) in enumerate(commands, 3):
        ws2.cell(row=row, column=1, value=name).font = Font(name='Calibri', bold=True, size=10)
        ws2.cell(row=row, column=2, value=cmd).font = Font(name='Courier New', size=9, color="1A237E")
        for col in range(1, 3):
            ws2.cell(row=row, column=col).border = make_border()
            ws2.cell(row=row, column=col).fill = make_fill(COLORS["row_alt"] if row % 2 == 0 else COLORS["row_main"])

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 80

    wb.save(filename)
    print(f"[OK] Generated: {filename}")


if __name__ == "__main__":
    base = os.path.dirname(os.path.abspath(__file__))
    print("=" * 60)
    print("  Scam Shield AI -- Security Report Generator")
    print("=" * 60)

    generate_findings_xlsx(os.path.join(base, "findings.xlsx"))
    generate_endpoint_xlsx(os.path.join(base, "endpoint-inventory.xlsx"))
    generate_dependency_xlsx(os.path.join(base, "dependency-report.xlsx"))

    print("\n[DONE] All 3 Excel security reports generated successfully!")
    print(f"  findings.xlsx            -- 41 security findings with severity, CWE, OWASP, fix")
    print(f"  endpoint-inventory.xlsx  -- 49 API endpoints with auth status and risk")
    print(f"  dependency-report.xlsx   -- {len(DEPS)} packages with CVE analysis")
